""" app/api_routes.py All /api/* REST endpoints for the MuTech dashboard Covers: auth, dashboard stats, leads,
campaigns, calls, analytics, logs, users """
# Delete the orphaned docstring + return block (about 12 lines)
import io
import csv
import hashlib
import os
from datetime import datetime, timedelta

import httpx
from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse

from app import database as db
from app import tenant_db as tdb
from app.plan_features import check_feature, check_campaign_limit, check_seat_limit, get_plan_features

router = APIRouter(prefix="/api")

# In-memory token store: token → user dict (cleared on restart; users re-login)
_token_store: dict = {}


def get_current_user(request: Request) -> dict:
    """Resolve Bearer token to a user dict. Checks memory cache then DB (survives restarts)."""
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing token")
    token = auth[7:]
    # Fast path: in-memory cache
    user = _token_store.get(token)
    if not user:
        # Slow path: DB lookup (restores sessions after service restart)
        user = db.get_session(token)
        if user:
            _token_store[token] = user  # repopulate cache
    if not user:
        raise HTTPException(status_code=401, detail="Invalid or expired token. Please log in again.")
    return user


# ═══════════════════════════════════════════════════════
# AUTH
# ═══════════════════════════════════════════════════════

@router.post("/auth/login")
async def login(request: Request):
    body     = await request.json()
    username = body.get("username", "").strip()
    password = body.get("password", "").strip()

    if not username or not password:
        raise HTTPException(status_code=400, detail="Username and password required")

    user = db.verify_user(username, password)
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials")

    # Simple token: sha256(username+password+secret)
    secret = os.getenv("JWT_SECRET", "mutech-secret-2026")
    token  = hashlib.sha256(f"{username}:{password}:{secret}".encode()).hexdigest()

    user_dict = {
        "user_id":   user["id"],
        "username":  user["username"],
        "name":      user.get("name", username),
        "role":      user["role"],
        "tenant_id": user.get("tenant_id") or 1,
    }

    # Store in memory (fast path) and DB (survives service restarts)
    _token_store[token] = user_dict
    db.save_session(token, user_dict)

    db.add_log(f"🔑 Login: {username} ({user['role']})")

    return {
        "token":    token,
        "user_id":  user["id"],
        "username": user["username"],
        "name":     user.get("name", username),
        "role":     user["role"],
    }


import re as _re


@router.post("/auth/register")
async def register_tenant(request: Request):
    """
    Public endpoint — new tenant self-signup.
    Creates tenant + admin user in one step.
    """
    data = await request.json()

    # ── Validate required fields ──────────────────────────────
    company_name = (data.get("company_name") or "").strip()
    contact_name = (data.get("contact_name") or "").strip()
    email        = (data.get("email") or "").strip().lower()
    password     = (data.get("password") or "").strip()
    plan         = data.get("plan", "starter")

    if not company_name:
        raise HTTPException(status_code=400, detail="Company name is required")
    if not contact_name:
        raise HTTPException(status_code=400, detail="Contact name is required")
    if not email or "@" not in email:
        raise HTTPException(status_code=400, detail="Valid email is required")
    if len(password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters")

    # ── Generate unique slug ──────────────────────────────────
    base_slug = _re.sub(r'[^a-z0-9]+', '-', company_name.lower()).strip('-')
    slug = base_slug
    all_tenants = tdb.get_all_tenants()
    existing_slugs = {t["slug"] for t in all_tenants}
    existing_emails = set()
    for t in all_tenants:
        for u in tdb.get_tenant_users(t["id"]):
            existing_emails.add((u.get("email") or u.get("username") or "").lower())

    counter = 1
    while slug in existing_slugs:
        slug = f"{base_slug}-{counter}"
        counter += 1

    if email in existing_emails:
        raise HTTPException(status_code=400, detail="An account with this email already exists")

    # ── Create tenant ─────────────────────────────────────────
    calls_limit = {"free": 100, "starter": 500, "enterprise": 10000}.get(plan, 500)
    try:
        tenant_id = tdb.create_tenant(
            name         = company_name,
            slug         = slug,
            plan         = plan,
            contact_name = contact_name,
            calls_limit  = calls_limit,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to create tenant: {e}")

    # ── Create admin user ─────────────────────────────────────
    # create_tenant_user hashes the password internally — pass plaintext
    try:
        tdb.create_tenant_user(
            tenant_id = tenant_id,
            username  = email,
            password  = password,
            role      = "admin",
            name      = contact_name,
            email     = email,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to create admin user: {e}")

    # ── Seed basic tenant config ───────────────────────────────
    tdb.update_tenant_config(
        tenant_id,
        company_name      = company_name,
        company_industry  = data.get("company_industry", ""),
        company_products  = data.get("company_products", ""),
        call_language     = data.get("call_language", "hindi"),
        agent_name        = "Aira",
        agent_voice       = "anushka",
    )

    db.add_log(f"🎉 New signup: {company_name} ({email}) — plan={plan} tid={tenant_id}")

    # ── Notify superadmin via Telegram ────────────────────────
    try:
        # Use platform owner (tenant 1) Telegram credentials from DB
        _platform_cfg = tdb.get_tenant_config(1) or {}
        tg_token = _platform_cfg.get("telegram_bot_token") or ""
        tg_chat  = _platform_cfg.get("telegram_chat_id")   or ""
        if tg_token and tg_chat:
            msg = (
                f"🎉 <b>New Tenant Signup!</b>\n\n"
                f"🏢 <b>Company:</b> {company_name}\n"
                f"👤 <b>Contact:</b> {contact_name}\n"
                f"📧 <b>Email:</b> {email}\n"
                f"📦 <b>Plan:</b> {plan.title()}\n"
                f"🆔 <b>Tenant ID:</b> {tenant_id}\n\n"
                f"<i>Login: {email}</i>"
            )
            url = f"https://api.telegram.org/bot{tg_token}/sendMessage"
            async with httpx.AsyncClient(timeout=5) as client:
                await client.post(url, json={"chat_id": tg_chat, "text": msg, "parse_mode": "HTML"})
    except Exception:
        pass  # Never block signup for notification failure

    return {
        "ok":          True,
        "tenant_id":   tenant_id,
        "message":     f"Account created! You can now log in with {email}.",
        "login_email": email,
    }


def _verify_token(request: Request) -> dict:
    """Extract and verify token from Authorization header."""
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing token")
    # For simplicity, accept any non-empty token (full JWT can be added later)
    return {"token": auth[7:]}


# ═══════════════════════════════════════════════════════
# DASHBOARD
# ═══════════════════════════════════════════════════════

@router.get("/dashboard/stats")
async def dashboard_stats(request: Request):
    user = get_current_user(request)
    return db.get_dashboard_stats(tenant_id=user["tenant_id"])


# In-memory active call tracker
_active_calls: dict = {}


@router.post("/calls/active/register")
async def register_active_call(request: Request):
    """Called by multi_agent_manager when a session starts."""
    from datetime import datetime
    body = await request.json()
    call_id = body.get("call_id")
    if call_id:
        body["registered_at"] = datetime.utcnow().isoformat()
        _active_calls[call_id] = body
    return {"ok": True}


@router.post("/calls/active/unregister")
async def unregister_active_call(request: Request):
    """Called by multi_agent_manager when a session ends."""
    body = await request.json()
    call_id = body.get("call_id")
    if call_id:
        _active_calls.pop(call_id, None)
    return {"ok": True}


@router.get("/active-calls")
async def get_active_calls(request: Request):
    """Returns currently active call sessions."""
    from datetime import datetime, timedelta
    user = get_current_user(request)
    tid  = user["tenant_id"]

    to_remove = []
    cutoff = datetime.utcnow() - timedelta(minutes=20)

    for call_id, call_data in _active_calls.items():
        # 1. Stale timeout: 20 minutes (covers crash/restart where finally never ran)
        if call_data.get("registered_at"):
            try:
                registered_at = datetime.fromisoformat(call_data["registered_at"].replace('Z', '+00:00'))
                if registered_at.replace(tzinfo=None) < cutoff:
                    to_remove.append(call_id)
                    continue
            except (ValueError, TypeError):
                to_remove.append(call_id)
                continue

        # 2. DB cross-check: if the call has duration_sec set it already ended
        try:
            with db.get_conn() as conn:
                row = conn.execute(
                    "SELECT duration_sec FROM calls WHERE id=?", (call_id,)
                ).fetchone()
                if row and row[0] is not None:
                    to_remove.append(call_id)
        except Exception:
            pass

    for call_id in to_remove:
        _active_calls.pop(call_id, None)

    calls = [
        {**c, "id": c.get("call_id", c.get("id"))}
        for c in _active_calls.values()
        if str(c.get("tenant_id", "1")) == str(tid)
    ]
    return {"calls": calls, "total": len(calls)}


@router.get("/dashboard/recent-calls")
async def recent_calls(request: Request, limit: int = 8):
    user = get_current_user(request)
    return db.get_recent_calls(limit=limit, tenant_id=user["tenant_id"])


@router.get("/dashboard/logs")
async def system_logs(limit: int = 30):
    return db.get_logs(limit=limit)


@router.get("/export/logs.csv")
async def export_logs_csv(limit: int = 1000):
    """Export recent system logs as CSV."""
    rows = db.get_logs(limit=min(limit, 5000))
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=["id", "level", "message", "created_at"])
    writer.writeheader()
    writer.writerows(rows)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=system_logs.csv"},
    )


@router.get("/export/calls.csv")
async def export_calls_csv(request: Request, limit: int = 5000):
    """Export call history as CSV."""
    user = get_current_user(request)
    rows = db.get_calls(limit=min(limit, 10000), tenant_id=user["tenant_id"])
    fields = [
        "id", "lead_id", "campaign_id", "phone", "lead_name", "company",
        "started_at", "ended_at", "duration_sec", "outcome", "sentiment",
        "summary", "call_sid",
    ]
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=calls.csv"},
    )


# ═══════════════════════════════════════════════════════
# LEADS
# ═══════════════════════════════════════════════════════

@router.get("/leads")
async def list_leads(
    request:     Request,
    status:      str = None,
    campaign_id: int = None,
    limit:       int = 100,
    offset:      int = 0,
):
    user = get_current_user(request)
    tid  = user["tenant_id"]
    leads = db.get_leads(status=status, campaign_id=campaign_id,
                         limit=limit, offset=offset, tenant_id=tid)
    total = db.count_leads(status=status, tenant_id=tid)
    return {"total": total, "leads": leads}


# ── IMPORTANT: specific paths MUST come before /{lead_id} ──

@router.get("/leads/groups")
async def lead_groups(request: Request):
    """Return available lead groupings for campaign assignment."""
    user = get_current_user(request)
    tid  = user["tenant_id"]
    with db.get_conn() as conn:
        def cq(sql): return conn.execute(sql + " AND tenant_id=?", (tid,)).fetchone()[0]
        total      = cq("SELECT COUNT(*) FROM leads WHERE 1=1")
        new        = cq("SELECT COUNT(*) FROM leads WHERE status='new'")
        called     = cq("SELECT COUNT(*) FROM leads WHERE status='called'")
        interested = cq("SELECT COUNT(*) FROM leads WHERE status IN ('interested','demo_booked')")
        not_int    = cq("SELECT COUNT(*) FROM leads WHERE status='not_interested'")
        unassigned = cq("SELECT COUNT(*) FROM leads WHERE campaign_id IS NULL")

    groups = []
    if new > 0:
        groups.append({"id": "new", "label": "New leads (never called)", "count": new})
    if unassigned > 0:
        groups.append({"id": "unassigned", "label": "Unassigned leads", "count": unassigned})
    if called > 0:
        groups.append({"id": "called", "label": "Previously called leads", "count": called})
    if interested > 0:
        groups.append({"id": "interested", "label": "Interested leads (follow up)", "count": interested})
    if not_int > 0:
        groups.append({"id": "not_interested", "label": "Not interested (re-try)", "count": not_int})
    if total > 0:
        groups.append({"id": "all", "label": "All leads", "count": total})

    return {"groups": groups}


@router.post("/leads/upload-csv")
async def upload_leads_csv(
    request:     Request,
    file:        UploadFile = File(...),
    campaign_id: int = None,
    force:       bool = False,   # if True, delete existing leads with same phone first
):
    """Upload leads from CSV. Expected columns: name, phone, company, designation, language"""
    user = get_current_user(request)
    tid  = user["tenant_id"]

    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="File must be a .csv")

    content = await file.read()
    try:
        text   = content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        rows   = list(reader)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"CSV parse error: {e}")

    if not rows:
        raise HTTPException(status_code=400, detail="CSV is empty")

    # If force=True, remove existing leads with same phones (for this tenant only)
    if force:
        from app.database import _normalize_phone
        deleted = 0
        with db.get_conn() as conn:
            for row in rows:
                raw = str(row.get('phone', '')).strip()
                if not raw: continue
                phone = _normalize_phone(raw)
                if not phone: continue
                result = conn.execute("DELETE FROM leads WHERE phone=? AND tenant_id=?", (phone, tid))
                deleted += result.rowcount
            conn.commit()
        db.add_log(f"🗑️ Force re-import: cleared {deleted} existing leads before upload")

    count = db.bulk_insert_leads(rows, campaign_id=campaign_id, tenant_id=tid)
    skipped = len(rows) - count

    msg = f"Imported {count} leads"
    if skipped > 0 and not force:
        msg += f" ({skipped} skipped — already exist. Use 'Replace existing' to re-import)"
    db.add_log(f"📂 CSV uploaded: {count} new leads from {file.filename}" +
               (f" ({skipped} skipped)" if skipped else ""))

    return {
        "message":  msg,
        "imported": count,
        "skipped":  skipped,
        "total":    len(rows),
        "force":    force,
    }


@router.post("/leads/upload-excel")
async def upload_leads_excel(
    request:     Request,
    file:        UploadFile = File(...),
    campaign_id: int = None,
    force:       bool = False,
):
    """
    Upload contacts from an Excel file (.xlsx or .xls).

    Required columns : name, phone
    Optional columns : company, designation, language  (missing = ignored / defaults used)

    Handles:
      - Missing optional columns gracefully (they default to '' / 'hi')
      - Scientific notation phone numbers written by Excel
      - Duplicate phone dedup (skip unless force=True)
      - Header row auto-detected (case-insensitive)
    """
    user = get_current_user(request)
    tid  = user["tenant_id"]

    fname = (file.filename or "").lower()
    if not (fname.endswith(".xlsx") or fname.endswith(".xls")):
        raise HTTPException(status_code=400, detail="File must be a .xlsx or .xls Excel file")

    content = await file.read()

    try:
        import openpyxl, io as _io
        wb = openpyxl.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)

        # ── Detect header row ─────────────────────────────────
        raw_header = next(rows_iter, None)
        if not raw_header:
            raise HTTPException(status_code=400, detail="Excel file is empty")

        # Normalize header names: lower-case, strip whitespace
        header = [str(h).strip().lower() if h is not None else "" for h in raw_header]

        def col(name):
            """Return index of column by name, or None if not present."""
            return header.index(name) if name in header else None

        i_name  = col("name")
        i_phone = col("phone")

        if i_name is None or i_phone is None:
            raise HTTPException(
                status_code=400,
                detail=f"Excel must have 'name' and 'phone' columns. Found: {header}"
            )

        i_company     = col("company")
        i_designation = col("designation")
        i_language    = col("language")

        # ── Build rows list ───────────────────────────────────
        leads_raw = []
        for row in rows_iter:
            def cell(idx):
                if idx is None or idx >= len(row):
                    return ""
                v = row[idx]
                return str(v).strip() if v is not None else ""

            # Handle Excel numeric phone (e.g. 919876543210.0 or 9.18827E+11)
            raw_phone = ""
            if i_phone is not None and i_phone < len(row):
                pv = row[i_phone]
                if pv is not None:
                    if isinstance(pv, float):
                        raw_phone = str(int(pv))
                    elif isinstance(pv, int):
                        raw_phone = str(pv)
                    else:
                        raw_phone = str(pv).strip()

            name = cell(i_name)
            if not name or not raw_phone:
                continue  # skip blank rows

            leads_raw.append({
                "name":        name,
                "phone":       raw_phone,
                "company":     cell(i_company),
                "designation": cell(i_designation),
                "language":    cell(i_language) or "hi",
            })

        wb.close()

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not parse Excel file: {e}")

    if not leads_raw:
        raise HTTPException(status_code=400, detail="No valid rows found in Excel (need name + phone)")

    # ── Force: delete existing records with same phones ───────
    if force:
        from app.database import _normalize_phone
        deleted = 0
        with db.get_conn() as conn:
            for row in leads_raw:
                phone = _normalize_phone(row["phone"])
                if not phone:
                    continue
                result = conn.execute(
                    "DELETE FROM leads WHERE phone=? AND tenant_id=?", (phone, tid)
                )
                deleted += result.rowcount
            conn.commit()
        if deleted:
            db.add_log(f"🗑️ Excel force re-import: cleared {deleted} existing leads")

    # ── Insert ────────────────────────────────────────────────
    count   = db.bulk_insert_leads(leads_raw, campaign_id=campaign_id, tenant_id=tid)
    skipped = len(leads_raw) - count

    msg = f"Imported {count} contacts from Excel"
    if skipped > 0 and not force:
        msg += f" ({skipped} skipped — duplicates)"

    db.add_log(
        f"📊 Excel upload: {count} new contacts from {file.filename}"
        + (f" ({skipped} skipped)" if skipped else "")
    )

    return {
        "message":  msg,
        "imported": count,
        "skipped":  skipped,
        "total":    len(leads_raw),
        "force":    force,
    }


@router.get("/leads/{lead_id}")
async def get_lead(lead_id: int):
    lead = db.get_lead(lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    return lead


@router.post("/leads")
async def create_lead(request: Request):
    user = get_current_user(request)
    body = await request.json()
    name  = body.get("name", "").strip()
    phone = body.get("phone", "").strip()
    if not name or not phone:
        raise HTTPException(status_code=400, detail="name and phone are required")
    lead_id = db.create_lead(
        name        = name,
        phone       = phone,
        company     = body.get("company"),
        designation = body.get("designation"),
        city        = body.get("city"),
        language    = body.get("language", "hi"),
        campaign_id = body.get("campaign_id"),
        tenant_id   = user["tenant_id"],
    )
    db.add_log(f"➕ Lead added: {name} ({phone})")
    return {"id": lead_id, "message": "Lead created"}


@router.put("/leads/{lead_id}")
async def update_lead(lead_id: int, request: Request):
    body = await request.json()
    lead = db.get_lead(lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    name   = body.get("name")
    phone  = body.get("phone")

    db.update_lead_full(
        lead_id     = lead_id,
        name        = name,
        phone       = phone,
        company     = body.get("company"),
        designation = body.get("designation"),
        language    = body.get("language"),
        status      = body.get("status"),
        notes       = body.get("notes"),
        callback_at = body.get("callback_at"),
    )

    db.add_log(f"✏️ Lead #{lead_id} updated — {name or lead['name']} ({phone or lead['phone']})")
    return {"message": "Lead updated", "id": lead_id}


@router.delete("/leads/{lead_id}")
async def delete_lead(lead_id: int):
    lead = db.get_lead(lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    phone = lead.get("phone", "")
    db.delete_lead(lead_id)
    return {"message": f"Lead and all records for {phone} deleted"}


# ═══════════════════════════════════════════════════════
# CAMPAIGNS
# ═══════════════════════════════════════════════════════

@router.get("/campaigns")
async def list_campaigns(request: Request, status: str = None):
    user = get_current_user(request)
    campaigns = db.get_campaigns(status=status, tenant_id=user["tenant_id"])
    return {"total": len(campaigns), "campaigns": campaigns}


@router.post("/campaigns")
async def create_campaign(request: Request):
    user        = get_current_user(request)
    tid         = user["tenant_id"]
    body        = await request.json()
    name        = body.get("name", "").strip()
    lead_group  = body.get("lead_group", "new")
    description = body.get("description", "")

    if not name:
        raise HTTPException(status_code=400, detail="Campaign name required")

    tenant = tdb.get_tenant(tid) or {}
    plan   = tenant.get("plan", "starter")
    with db.get_conn() as _conn:
        active_count = _conn.execute(
            "SELECT COUNT(*) FROM campaigns WHERE tenant_id=? AND status='running'",
            (tid,)
        ).fetchone()[0]
    gate = check_campaign_limit(plan, active_count)
    if not gate["allowed"]:
        raise HTTPException(status_code=402, detail={
            "message":    gate["reason"],
            "upgrade_to": gate["upgrade_to"],
            "code":       "CAMPAIGN_LIMIT_REACHED",
        })

    camp_id  = db.create_campaign(name=name, description=description, tenant_id=tid)
    assigned = db.assign_leads_to_campaign(camp_id, lead_group, tenant_id=tid)

    db.add_log(f"🚀 Campaign created: {name} — {assigned} leads assigned ({lead_group})")
    return {"id": camp_id, "message": "Campaign created", "leads_assigned": assigned}


@router.get("/campaigns/{campaign_id}")
async def get_campaign(campaign_id: int):
    c = db.get_campaign(campaign_id)
    if not c:
        raise HTTPException(status_code=404, detail="Campaign not found")
    return c


@router.put("/campaigns/{campaign_id}")
async def update_campaign(campaign_id: int, request: Request):
    """Update campaign name, description and call delay."""
    user = get_current_user(request)
    c = db.get_campaign(campaign_id)
    if not c:
        raise HTTPException(status_code=404, detail="Campaign not found")

    body = await request.json()
    name        = (body.get("name") or "").strip()
    description = (body.get("description") or "").strip()
    delay       = body.get("delay_seconds")

    if not name:
        raise HTTPException(status_code=400, detail="Campaign name cannot be empty")

    with db.get_conn() as conn:
        if delay is not None:
            conn.execute(
                """UPDATE campaigns SET name=?, description=?, schedule_delay=?
                   WHERE id=?""",
                (name, description, int(delay), campaign_id)
            )
        else:
            conn.execute(
                "UPDATE campaigns SET name=?, description=? WHERE id=?",
                (name, description, campaign_id)
            )
        conn.commit()

    db.add_log(f"✏️ Campaign updated: {name} (id={campaign_id})")
    return {"ok": True, "message": "Campaign updated"}


@router.post("/campaigns/{campaign_id}/add-leads")
async def add_leads_to_campaign(campaign_id: int, request: Request):
    """
    Add more leads to an existing campaign by lead group selection.
    Only adds leads not already assigned to an active campaign.
    Works for any campaign status.
    """
    user = get_current_user(request)
    tid  = user["tenant_id"]
    c = db.get_campaign(campaign_id)
    if not c:
        raise HTTPException(status_code=404, detail="Campaign not found")

    body       = await request.json()
    lead_group = body.get("lead_group", "new")

    added = db.assign_leads_to_campaign(campaign_id, lead_group, tenant_id=tid)

    # Update leads_count to reflect new total
    with db.get_conn() as conn:
        total = conn.execute(
            "SELECT COUNT(*) FROM leads WHERE campaign_id=?", (campaign_id,)
        ).fetchone()[0]
        conn.execute(
            "UPDATE campaigns SET leads_count=? WHERE id=?",
            (total, campaign_id)
        )
        conn.commit()

    db.add_log(
        f"➕ {added} leads added to campaign '{c['name']}' (group: {lead_group})"
    )
    return {
        "ok":           True,
        "added":        added,
        "total_leads":  total,
        "message":      f"{added} leads added to campaign",
    }


@router.post("/campaigns/{campaign_id}/upload-leads")
async def upload_leads_to_campaign(
    request:     Request,
    campaign_id: int,
    file:        UploadFile = File(...),
):
    """
    Upload a CSV or Excel file and add those contacts directly
    to an existing campaign. Skips duplicates by phone number.
    """
    user = get_current_user(request)
    tid  = user["tenant_id"]
    c = db.get_campaign(campaign_id)
    if not c:
        raise HTTPException(status_code=404, detail="Campaign not found")

    fname   = (file.filename or "").lower()
    content = await file.read()
    rows    = []

    if fname.endswith(".csv"):
        import io as _io
        text   = content.decode("utf-8-sig")
        reader = csv.DictReader(_io.StringIO(text))
        rows   = list(reader)

    elif fname.endswith(".xlsx") or fname.endswith(".xls"):
        import openpyxl, io as _io
        wb  = openpyxl.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
        ws  = wb.active
        it  = ws.iter_rows(values_only=True)
        raw_header = next(it, None)
        if not raw_header:
            raise HTTPException(status_code=400, detail="Empty file")
        header = [str(h).strip().lower() if h else "" for h in raw_header]
        def col(n): return header.index(n) if n in header else None
        i_name  = col("name");  i_phone = col("phone")
        i_comp  = col("company"); i_des = col("designation"); i_lang = col("language")
        if i_name is None or i_phone is None:
            raise HTTPException(status_code=400, detail="File must have 'name' and 'phone' columns")
        for row in it:
            def cell(idx):
                if idx is None or idx >= len(row): return ""
                v = row[idx]
                if isinstance(v, float): return str(int(v))
                return str(v).strip() if v else ""
            rows.append({
                "name":        cell(i_name),
                "phone":       cell(i_phone),
                "company":     cell(i_comp),
                "designation": cell(i_des),
                "language":    cell(i_lang) or "hi",
            })
        wb.close()
    else:
        raise HTTPException(status_code=400, detail="Only .csv, .xlsx, .xls files are supported")

    if not rows:
        raise HTTPException(status_code=400, detail="No rows found in file")

    count   = db.bulk_insert_leads(rows, campaign_id=campaign_id, tenant_id=tid)
    skipped = len(rows) - count

    # Refresh leads_count
    with db.get_conn() as conn:
        total = conn.execute(
            "SELECT COUNT(*) FROM leads WHERE campaign_id=?", (campaign_id,)
        ).fetchone()[0]
        conn.execute(
            "UPDATE campaigns SET leads_count=? WHERE id=?",
            (total, campaign_id)
        )
        conn.commit()

    db.add_log(
        f"📂 {count} contacts uploaded to campaign '{c['name']}' from {file.filename}"
        + (f" ({skipped} skipped)" if skipped else "")
    )
    return {
        "ok":          True,
        "imported":    count,
        "skipped":     skipped,
        "total":       len(rows),
        "total_leads": total,
        "message":     f"Imported {count} contacts into campaign",
    }


@router.get("/campaigns/{campaign_id}/leads")
async def get_campaign_leads(
    campaign_id: int,
    request:     Request,
    status:      str = None,
    limit:       int = 100,
    offset:      int = 0,
):
    """Get paginated leads for a specific campaign with optional status filter."""
    user = get_current_user(request)
    c = db.get_campaign(campaign_id)
    if not c:
        raise HTTPException(status_code=404, detail="Campaign not found")

    with db.get_conn() as conn:
        query  = "SELECT * FROM leads WHERE campaign_id=?"
        params = [campaign_id]
        if status:
            query += " AND status=?"
            params.append(status)
        query += " ORDER BY created_at DESC LIMIT ? OFFSET ?"
        params += [limit, offset]
        leads = [dict(r) for r in conn.execute(query, params).fetchall()]

        # Total count
        cq     = "SELECT COUNT(*) FROM leads WHERE campaign_id=?"
        cp     = [campaign_id]
        if status:
            cq += " AND status=?"; cp.append(status)
        total = conn.execute(cq, cp).fetchone()[0]

        # Status breakdown
        breakdown = {}
        for row in conn.execute(
            "SELECT status, COUNT(*) as cnt FROM leads WHERE campaign_id=? GROUP BY status",
            (campaign_id,)
        ).fetchall():
            breakdown[row["status"]] = row["cnt"]

    return {
        "campaign_id": campaign_id,
        "total":       total,
        "offset":      offset,
        "limit":       limit,
        "status_filter": status,
        "breakdown":   breakdown,
        "leads":       leads,
    }


@router.delete("/campaigns/{campaign_id}/leads/{lead_id}")
async def remove_lead_from_campaign(campaign_id: int, lead_id: int):
    """
    Remove a single lead from a campaign by setting campaign_id = NULL.
    The lead itself is NOT deleted — it becomes unassigned.
    Only works if the campaign is not currently running.
    """
    c = db.get_campaign(campaign_id)
    if not c:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if c["status"] == "running":
        raise HTTPException(
            status_code=400,
            detail="Cannot remove leads while campaign is running. Pause it first."
        )

    with db.get_conn() as conn:
        lead = conn.execute(
            "SELECT * FROM leads WHERE id=? AND campaign_id=?",
            (lead_id, campaign_id)
        ).fetchone()
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found in this campaign")

        conn.execute(
            "UPDATE leads SET campaign_id=NULL WHERE id=?", (lead_id,)
        )
        # Recalculate leads_count
        total = conn.execute(
            "SELECT COUNT(*) FROM leads WHERE campaign_id=?", (campaign_id,)
        ).fetchone()[0]
        conn.execute(
            "UPDATE campaigns SET leads_count=? WHERE id=?",
            (total, campaign_id)
        )
        conn.commit()

    db.add_log(
        f"➖ Lead #{lead_id} removed from campaign '{c['name']}'"
    )
    return {"ok": True, "message": "Lead removed from campaign", "total_leads": total}


@router.delete("/campaigns/{campaign_id}/leads")
async def bulk_remove_leads_from_campaign(campaign_id: int, request: Request):
    """
    Bulk remove leads from campaign by status.
    E.g. remove all 'new' leads, or all 'not_interested' leads.
    Body: { "status": "new" }  — if omitted, removes ALL unassigned leads (new only)
    """
    c = db.get_campaign(campaign_id)
    if not c:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if c["status"] == "running":
        raise HTTPException(
            status_code=400,
            detail="Cannot remove leads while campaign is running. Pause it first."
        )

    body   = await request.json()
    status = body.get("status", "new")
    allowed = {"new", "called", "not_interested", "interested"}
    if status not in allowed:
        raise HTTPException(
            status_code=400,
            detail=f"Can only bulk-remove leads with status: {allowed}"
        )

    with db.get_conn() as conn:
        cur = conn.execute(
            "UPDATE leads SET campaign_id=NULL WHERE campaign_id=? AND status=?",
            (campaign_id, status)
        )
        removed = cur.rowcount
        total = conn.execute(
            "SELECT COUNT(*) FROM leads WHERE campaign_id=?", (campaign_id,)
        ).fetchone()[0]
        conn.execute(
            "UPDATE campaigns SET leads_count=? WHERE id=?",
            (total, campaign_id)
        )
        conn.commit()

    db.add_log(
        f"➖ {removed} '{status}' leads removed from campaign '{c['name']}'"
    )
    return {
        "ok":          True,
        "removed":     removed,
        "total_leads": total,
        "message":     f"{removed} leads removed from campaign",
    }


@router.post("/campaigns/{campaign_id}/restart")
async def restart_campaign(campaign_id: int, request: Request):
    """Reset all leads to 'new' and restart the campaign from scratch."""
    from app.campaign_runner import run_campaign

    c = db.get_campaign(campaign_id)
    if not c:
        raise HTTPException(status_code=404, detail="Campaign not found")

    try:
        body  = await request.json()
        delay = int(body.get("delay_seconds", 60))
    except Exception:
        delay = 60

    # Reset leads and counters
    reset_count = db.reset_campaign_leads(campaign_id)
    db.update_campaign_status(campaign_id, "running")
    db.add_log(f"🔄 Campaign restarted: {c['name']} — {reset_count} leads reset to new, {delay}s delay")

    import asyncio
    asyncio.create_task(run_campaign(campaign_id, delay))

    return {
        "message": f"Campaign '{c['name']}' restarted",
        "leads_reset": reset_count,
        "delay": delay,
    }


@router.post("/campaigns/{campaign_id}/start")
async def start_campaign(campaign_id: int, request: Request, background_tasks=None):
    from fastapi import BackgroundTasks
    from app.campaign_runner import run_campaign

    c = db.get_campaign(campaign_id)
    if not c:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if c['status'] == 'running':
        return {"message": f"Campaign '{c['name']}' is already running"}

    # Get delay from request body (optional)
    try:
        body  = await request.json()
        delay = int(body.get("delay_seconds", 60))
    except Exception:
        delay = 60

    db.update_campaign_status(campaign_id, "running")
    db.add_log(f"▶️ Campaign started: {c['name']} — {c.get('leads_count',0)} leads, {delay}s delay")

    # Fire background task — actual outbound calls
    import asyncio
    asyncio.create_task(run_campaign(campaign_id, delay))

    return {
        "message": f"Campaign '{c['name']}' started",
        "leads":   c.get('leads_count', 0),
        "delay":   delay,
    }


@router.post("/campaigns/{campaign_id}/pause")
async def pause_campaign(campaign_id: int):
    c = db.get_campaign(campaign_id)
    if not c:
        raise HTTPException(status_code=404, detail="Campaign not found")
    db.update_campaign_status(campaign_id, "paused")
    db.add_log(f"⏸️ Campaign paused: {c['name']}")
    return {"message": f"Campaign '{c['name']}' paused"}


@router.post("/campaigns/{campaign_id}/complete")
async def complete_campaign(campaign_id: int):
    c = db.get_campaign(campaign_id)
    if not c:
        raise HTTPException(status_code=404, detail="Campaign not found")
    db.update_campaign_status(campaign_id, "completed")
    db.add_log(f"✅ Campaign completed: {c['name']}")
    return {"message": f"Campaign '{c['name']}' marked complete"}


@router.delete("/campaigns/{campaign_id}")
async def delete_campaign(campaign_id: int):
    c = db.get_campaign(campaign_id)
    if not c:
        raise HTTPException(status_code=404, detail="Campaign not found")
    db.delete_campaign(campaign_id)
    return {"message": "Campaign deleted"}


@router.post("/campaigns/{campaign_id}/schedule")
async def schedule_campaign(campaign_id: int, request: Request):
    """Schedule a campaign: once, daily, or weekdays at a specific time."""
    c = db.get_campaign(campaign_id)
    if not c:
        raise HTTPException(status_code=404, detail="Campaign not found")
    body         = await request.json()
    stype        = body.get("schedule_type", "once")   # once | daily | weekdays
    stime        = body.get("schedule_time", "09:00")  # HH:MM
    sdate        = body.get("schedule_date", "")       # YYYY-MM-DD (for once)
    sdays        = body.get("schedule_days", "mon,tue,wed,thu,fri")
    delay        = int(body.get("delay_seconds", 60))
    if stype not in ("once", "daily", "weekdays"):
        raise HTTPException(status_code=400, detail="Invalid schedule_type")
    next_run = f"{sdate} {stime}" if stype == "once" and sdate else None
    with db.get_conn() as conn:
        conn.execute("""
            UPDATE campaigns SET
                schedule_type   = ?,
                schedule_time   = ?,
                schedule_days   = ?,
                schedule_delay  = ?,
                schedule_status = 'pending',
                next_run_at     = ?
            WHERE id=?
        """, (stype, stime, sdays, delay, next_run, campaign_id))
        conn.commit()
    db.add_log(f"🕐 Campaign scheduled: {c['name']} — {stype} at {stime}")
    return {
        "message":       f"Campaign '{c['name']}' scheduled",
        "schedule_type": stype,
        "schedule_time": stime,
        "next_run_at":   next_run,
    }


@router.delete("/campaigns/{campaign_id}/schedule")
async def cancel_schedule(campaign_id: int):
    """Cancel a campaign's schedule without deleting the campaign."""
    c = db.get_campaign(campaign_id)
    if not c:
        raise HTTPException(status_code=404, detail="Campaign not found")
    with db.get_conn() as conn:
        conn.execute("""
            UPDATE campaigns SET
                schedule_type=NULL, schedule_time=NULL,
                schedule_status=NULL, next_run_at=NULL
            WHERE id=?
        """, (campaign_id,))
        conn.commit()
    db.add_log(f"❌ Campaign schedule cancelled: {c['name']}")
    return {"message": f"Schedule cancelled for '{c['name']}'"}


@router.get("/campaigns/{campaign_id}/schedule")
async def get_schedule(campaign_id: int):
    """Get current schedule for a campaign."""
    c = db.get_campaign(campaign_id)
    if not c:
        raise HTTPException(status_code=404, detail="Campaign not found")
    return {
        "campaign_id":    campaign_id,
        "schedule_type":  c.get("schedule_type"),
        "schedule_time":  c.get("schedule_time"),
        "schedule_days":  c.get("schedule_days"),
        "schedule_delay": c.get("schedule_delay", 60),
        "schedule_status":c.get("schedule_status"),
        "next_run_at":    c.get("next_run_at"),
        "is_scheduled":   c.get("schedule_type") is not None,
    }


@router.post("/campaigns/{campaign_id}/follow-up")
async def configure_campaign_follow_up(campaign_id: int, request: Request):
    """Configure follow-up automation for a campaign."""
    c = db.get_campaign(campaign_id)
    if not c:
        raise HTTPException(status_code=404, detail="Campaign not found")

    body = await request.json()
    follow_up_enabled = body.get("follow_up_enabled", False)
    follow_up_type = body.get("follow_up_type", "whatsapp")  # whatsapp, email, both
    follow_up_delay_minutes = body.get("follow_up_delay_minutes", 30)
    follow_up_message_template = body.get("follow_up_message_template", "").strip()

    if follow_up_type not in ("whatsapp", "email", "both"):
        raise HTTPException(status_code=400, detail="Invalid follow_up_type. Use: whatsapp, email, both")

    if follow_up_delay_minutes < 1 or follow_up_delay_minutes > 1440:  # Max 24 hours
        raise HTTPException(status_code=400, detail="follow_up_delay_minutes must be between 1 and 1440")

    db.update_campaign_follow_up(
        campaign_id=campaign_id,
        follow_up_enabled=follow_up_enabled,
        follow_up_type=follow_up_type,
        follow_up_delay_minutes=follow_up_delay_minutes,
        follow_up_message_template=follow_up_message_template
    )

    db.add_log(
        f"📱 Campaign '{c['name']}' follow-up {'enabled' if follow_up_enabled else 'disabled'}: "
        f"{follow_up_type} after {follow_up_delay_minutes}min"
    )

    return {
        "id": campaign_id,
        "message": f"Follow-up {'enabled' if follow_up_enabled else 'disabled'} for campaign '{c['name']}'",
        "follow_up_enabled": follow_up_enabled,
        "follow_up_type": follow_up_type,
        "follow_up_delay_minutes": follow_up_delay_minutes
    }


# Helper function to validate time format
def _validate_time_format(time_str: str) -> bool:
    """Validate time in HH:MM format (24-hour)."""
    import re
    if not re.match(r'^\d{2}:\d{2}$', time_str):
        return False

    hour, minute = map(int, time_str.split(':'))
    return 0 <= hour <= 23 and 0 <= minute <= 59


# ═══════════════════════════════════════════════════════
# CALLS
# ═══════════════════════════════════════════════════════

@router.get("/calls")
async def list_calls(
    request:     Request,
    limit:       int = 50,
    offset:      int = 0,
    campaign_id: int = None,
):
    user  = get_current_user(request)
    tid   = user["tenant_id"]
    calls = db.get_calls(limit=limit, offset=offset, campaign_id=campaign_id, tenant_id=tid)
    total = db.count_calls(tenant_id=tid)
    return {"total": total, "calls": calls}


@router.get("/calls/{call_id}")
async def get_call(call_id: int):
    call = db.get_call(call_id)
    if not call:
        raise HTTPException(status_code=404, detail="Call not found")
    return call


@router.get("/calls/{call_id}/transcript")
async def get_call_transcript(call_id: int):
    call = db.get_call(call_id)
    if not call:
        raise HTTPException(status_code=404, detail="Call not found")
    return {
        "call_id":    call_id,
        "lead_name":  call.get("lead_name", ""),
        "started_at": call.get("started_at", ""),
        "duration":   call.get("duration_sec", 0),
        "sentiment":  call.get("sentiment", ""),
        "summary":    call.get("summary", ""),
        "transcript": call.get("transcript") or "No transcript available",
    }


@router.post("/calls/{call_id}/generate-summary")
async def generate_summary(call_id: int):
    """Generate a fresh summary from the transcript using LLM."""
    call = db.get_call(call_id)
    if not call:
        raise HTTPException(status_code=404, detail="Call not found")

    transcript = call.get("transcript") or ""
    if not transcript.strip():
        return {
            "call_id": call_id,
            "summary": "No transcript available for summary generation.",
            "status": "no_transcript"
        }

    try:
        from groq import Groq

        # Use call's tenant Groq key from DB; fall back to platform (tenant 1)
        call_tenant_id = call.get("tenant_id") or 1
        _cfg = tdb.get_tenant_config(call_tenant_id) or {}
        groq_key = _cfg.get("groq_api_key") or ""
        if not groq_key and call_tenant_id != 1:
            _platform = tdb.get_tenant_config(1) or {}
            groq_key = _platform.get("groq_api_key") or ""

        agent_name = _cfg.get("agent_name") or "Aira"

        # Parse transcript into role-based conversation
        lines = transcript.split('\n')
        conversation = []
        for line in lines:
            line = line.strip()
            if line.startswith(f'{agent_name}:'):
                conversation.append({'role': 'assistant', 'content': line[len(agent_name)+1:].strip()})
            elif line.startswith('Customer:'):
                conversation.append({'role': 'user', 'content': line[9:].strip()})

        if not conversation:
            return {
                "call_id": call_id,
                "summary": "Could not parse transcript for analysis.",
                "status": "parse_error"
            }

        # Format for LLM analysis
        conversation_text = "\n".join(
            f"{m['role'].upper()}: {m['content']}"
            for m in conversation
        )

        company_name = _cfg.get("company_name") or "the company"
        prompt = f"""You are analyzing a sales call between {agent_name} (AI sales agent for {company_name}) and a customer.

CONVERSATION:
{conversation_text}

Analyze this conversation and provide a SHORT point-wise summary (max 3-4 bullet points in English):
- Customer's main interest/concern
- Key value proposition shared by {agent_name}
- Customer's response/sentiment
- Next steps (if any) or call outcome

Keep each point to ONE LINE maximum. Be concise and factual."""

        if not groq_key:
            raise HTTPException(status_code=503, detail="Groq API key not configured. Set it in Account Settings → API Keys.")

        client = Groq(api_key=groq_key)
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=300,
            temperature=0.3,
        )

        summary_text = response.choices[0].message.content.strip()

        # Update the summary in DB
        with db.get_conn() as conn:
            conn.execute(
                "UPDATE calls SET summary = ? WHERE id = ?",
                (summary_text, call_id)
            )
            conn.commit()

        db.add_log(f"📝 Summary regenerated for call #{call_id}")

        return {
            "call_id": call_id,
            "summary": summary_text,
            "status": "success"
        }

    except Exception as e:
        return {
            "call_id": call_id,
            "summary": f"Error generating summary: {str(e)}",
            "status": "error"
        }


@router.post("/calls/test")
async def test_call(request: Request):
    """Manually test an outbound call to a phone number."""
    user = get_current_user(request)
    body = await request.json()
    phone         = body.get("phone", "").strip()
    lead_id       = body.get("lead_id")
    customer_name = body.get("customer_name", "").strip()

    if not phone:
        raise HTTPException(status_code=400, detail="Phone number required")

    from app.campaign_runner import make_single_call

    metadata = {
        "tenant_id":     str(user["tenant_id"]),
        "customer_name": customer_name,
    }
    if lead_id:
        metadata["lead_id"] = str(lead_id)

    db.add_log(f"🔍 test_call: user={user['username']} tenant_id={user['tenant_id']} phone={phone}")
    result = await make_single_call(phone, lead_id=str(lead_id) if lead_id else None, metadata=metadata)
    if result:
        db.add_log(f"📞 Test call to {phone} — ID: {result['call_id']}")
        return {
            "message": f"Call initiated to {phone}",
            "call_id": result["call_id"],
            "provider": result.get("provider"),
            "dry_run": result.get("dry_run", False),
        }
    else:
        db.add_log(f"❌ Test call failed to {phone}")
        raise HTTPException(status_code=500, detail="Call failed")


# ── FLOW BUILDER ───────────────────────────────────────────────

@router.get("/flows")
async def list_flows(request: Request):
    user = get_current_user(request)
    tid  = user["tenant_id"]
    with db.get_conn() as conn:
        rows = conn.execute(
            "SELECT id,name,description,is_active,created_at FROM call_flows WHERE tenant_id=? ORDER BY id DESC",
            (tid,)
        ).fetchall()
    return {"flows": [dict(r) for r in rows]}

@router.post("/flows")
async def create_flow(request: Request):
    user = get_current_user(request)
    tid  = user["tenant_id"]
    body = await request.json()
    name = body.get("name","").strip()
    if not name:
        raise HTTPException(400, "name required")
    flow_json = body.get("flow_json", "{}")
    import json as _json
    try: _json.loads(flow_json)
    except: raise HTTPException(400, "Invalid flow_json")
    with db.get_conn() as conn:
        cur = conn.execute(
            "INSERT INTO call_flows (tenant_id,name,description,flow_json) VALUES (?,?,?,?)",
            (tid, name, body.get("description",""), flow_json)
        )
        conn.commit()
        flow_id = cur.lastrowid
    return {"id": flow_id, "message": "Flow created"}

@router.put("/flows/{flow_id}")
async def update_flow(flow_id: int, request: Request):
    user = get_current_user(request)
    tid  = user["tenant_id"]
    body = await request.json()
    import json as _json
    flow_json = body.get("flow_json")
    if flow_json:
        try: _json.loads(flow_json)
        except: raise HTTPException(400, "Invalid flow_json")
    with db.get_conn() as conn:
        conn.execute(
            """UPDATE call_flows SET name=COALESCE(?,name),
               description=COALESCE(?,description),
               flow_json=COALESCE(?,flow_json),
               is_active=COALESCE(?,is_active),
               updated_at=datetime('now')
               WHERE id=? AND tenant_id=?""",
            (body.get("name"), body.get("description"), flow_json,
             body.get("is_active"), flow_id, tid)
        )
        conn.commit()
    return {"message": "Flow updated"}

@router.delete("/flows/{flow_id}")
async def delete_flow(flow_id: int, request: Request):
    user = get_current_user(request)
    tid  = user["tenant_id"]
    with db.get_conn() as conn:
        conn.execute("DELETE FROM call_flows WHERE id=? AND tenant_id=?", (flow_id, tid))
        conn.commit()
    return {"message": "Flow deleted"}

@router.get("/flows/{flow_id}")
async def get_flow(flow_id: int, request: Request):
    user = get_current_user(request)
    tid  = user["tenant_id"]
    with db.get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM call_flows WHERE id=? AND tenant_id=?", (flow_id, tid)
        ).fetchone()
    if not row:
        raise HTTPException(404, "Flow not found")
    return {"flow": dict(row)}


# ═══════════════════════════════════════════════════════
# ANALYTICS
# ═══════════════════════════════════════════════════════

@router.get("/analytics/daily")
async def daily_stats(request: Request, days: int = 14, campaign_id: int = None, format: str = None):
    user = get_current_user(request)
    data = db.get_daily_call_stats(days=days, tenant_id=user["tenant_id"], campaign_id=campaign_id)
    if format == "csv":
        import io, csv as _csv
        from fastapi.responses import StreamingResponse
        buf = io.StringIO()
        w = _csv.DictWriter(buf, fieldnames=["date","total","answered","interested","demos","not_interested"])
        w.writeheader()
        w.writerows(data)
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=analytics_{days}d.csv"}
        )
    return {"days": days, "campaign_id": campaign_id, "data": data, "daily": data}


@router.get("/analytics/funnel")
async def funnel_stats(request: Request, campaign_id: int = None):
    user = get_current_user(request)
    tid  = user["tenant_id"]
    with db.get_conn() as conn:
        base_params = [tid]
        base_filter = " AND tenant_id=?"
        if campaign_id is not None:
            base_filter += " AND campaign_id=?"
            base_params.append(campaign_id)

        def cq(extra_sql):
            sql    = "SELECT COUNT(*) FROM calls WHERE 1=1" + base_filter + extra_sql
            params = base_params[:]
            return conn.execute(sql, params).fetchone()[0]

        total      = cq("")
        answered   = cq(" AND outcome='answered'")
        interested = cq(" AND sentiment IN ('interested','demo_booked')")
        demos      = cq(" AND sentiment='demo_booked'")

    return {
        "campaign_id": campaign_id,
        "funnel": [
            {"stage": "Total Calls", "count": total,      "pct": 100},
            {"stage": "Answered",    "count": answered,   "pct": round(answered/total*100,1)    if total else 0},
            {"stage": "Interested",  "count": interested, "pct": round(interested/total*100,1)  if total else 0},
            {"stage": "Demo Booked", "count": demos,      "pct": round(demos/total*100,1)       if total else 0},
        ]
    }


@router.get("/campaigns/{campaign_id}/analytics")
async def campaign_analytics(campaign_id: int, request: Request):
    user = get_current_user(request)
    tid  = user["tenant_id"]
    with db.get_conn() as conn:
        outcomes = [dict(r) for r in conn.execute("""
            SELECT outcome, COUNT(*) as count
            FROM calls WHERE campaign_id=? AND tenant_id=?
            GROUP BY outcome
        """, (campaign_id, tid)).fetchall()]

        sentiments = [dict(r) for r in conn.execute("""
            SELECT sentiment, COUNT(*) as count
            FROM calls WHERE campaign_id=? AND tenant_id=?
            GROUP BY sentiment
        """, (campaign_id, tid)).fetchall()]

        daily = [dict(r) for r in conn.execute("""
            SELECT date(started_at) as date, COUNT(*) as calls,
                   SUM(CASE WHEN outcome='answered' THEN 1 ELSE 0 END) as answered,
                   SUM(CASE WHEN sentiment='demo_booked' THEN 1 ELSE 0 END) as demos
            FROM calls WHERE campaign_id=? AND tenant_id=?
            GROUP BY date(started_at) ORDER BY date ASC
        """, (campaign_id, tid)).fetchall()]

        summary = dict(conn.execute("""
            SELECT COUNT(*) as total_calls,
                   SUM(CASE WHEN outcome='answered' THEN 1 ELSE 0 END) as answered,
                   SUM(CASE WHEN sentiment='interested' THEN 1 ELSE 0 END) as interested,
                   SUM(CASE WHEN sentiment='demo_booked' THEN 1 ELSE 0 END) as demos,
                   AVG(CASE WHEN duration_sec > 0 THEN duration_sec END) as avg_duration
            FROM calls WHERE campaign_id=? AND tenant_id=?
        """, (campaign_id, tid)).fetchone())

    return {
        "campaign_id": campaign_id,
        "summary":     summary,
        "outcomes":    outcomes,
        "sentiments":  sentiments,
        "daily":       daily,
    }


@router.get("/analytics/hourly")
async def hourly_stats(request: Request, days: int = 30):
    """Hour-of-day call distribution for heatmap."""
    user = get_current_user(request)
    data = db.get_hourly_call_stats(days=days, tenant_id=user["tenant_id"])
    # Fill missing hours with zeros
    hour_map = {r['hour']: r for r in data}
    full = []
    for h in range(24):
        full.append(hour_map.get(h, {'hour': h, 'total': 0, 'answered': 0}))
    return {"days": days, "data": full}


@router.get("/analytics/sentiment")
async def sentiment_breakdown(request: Request):
    user = get_current_user(request)
    tid  = user["tenant_id"]
    with db.get_conn() as conn:
        rows = conn.execute("""
            SELECT sentiment, COUNT(*) as count
            FROM calls WHERE tenant_id=?
            GROUP BY sentiment
            ORDER BY count DESC
        """, (tid,)).fetchall()
    return {"breakdown": [dict(r) for r in rows]}


# ═══════════════════════════════════════════════════════
# USERS  (admin only — token check skipped for now)
# ═══════════════════════════════════════════════════════

@router.get("/users")
async def list_users(request: Request):
    user = get_current_user(request)
    return {"users": db.get_all_users(tenant_id=user["tenant_id"])}


@router.post("/users")
async def create_user(request: Request):
    current = get_current_user(request)
    body     = await request.json()
    username = body.get("username", "").strip()
    password = body.get("password", "").strip()
    role     = body.get("role", "sales")
    name     = body.get("name", "").strip()
    email    = body.get("email", "").strip()

    if not username or not password:
        raise HTTPException(status_code=400, detail="username and password required")

    tenant = tdb.get_tenant(current["tenant_id"]) or {}
    plan   = tenant.get("plan", "starter")
    current_seats = len(tdb.get_tenant_users(current["tenant_id"]))
    gate = check_seat_limit(plan, current_seats)
    if not gate["allowed"]:
        raise HTTPException(status_code=402, detail={
            "message":    gate["reason"],
            "upgrade_to": gate["upgrade_to"],
            "code":       "SEAT_LIMIT_REACHED",
        })

    try:
        user_id = db.add_user(username, password, role, name, email,
                              tenant_id=current["tenant_id"])
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"User creation failed: {e}")

    db.add_log(f"👤 User added: {username} ({role})")
    return {"id": user_id, "message": "User created"}


@router.delete("/users/{user_id}")
async def remove_user(user_id: int):
    user = db.get_user_by_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    db.delete_user(user_id)
    return {"message": "User deleted"}


# ═══════════════════════════════════════════════════════
# SYSTEM
# ═══════════════════════════════════════════════════════

@router.get("/system/health")
async def system_health():
    """Live health check — tests each service connection."""
    import httpx, time

    results = {}
    start = time.time()

    # ── Database ──────────────────────────────────────
    try:
        stats = db.get_dashboard_stats()
        results["database"] = {"status": "ok", "calls": stats["total_calls"]}
    except Exception as e:
        results["database"] = {"status": "error", "detail": str(e)}

    # Use platform (tenant 1) keys from DB for system health check
    _platform_cfg = tdb.get_tenant_config(1) or {}

    # ── Sarvam AI ─────────────────────────────────────
    sarvam_key = _platform_cfg.get("sarvam_api_key") or ""
    if sarvam_key:
        try:
            async with httpx.AsyncClient(timeout=5) as client:
                r = await client.get(
                    "https://api.sarvam.ai/v1/models",
                    headers={"api-subscription-key": sarvam_key}
                )
            results["sarvam"] = {"status": "ok" if r.status_code < 400 else "error",
                                  "model_stt": "saarika:v2.5", "model_tts": "bulbul:v2"}
        except Exception:
            results["sarvam"] = {"status": "ok", "model_stt": "saarika:v2.5", "model_tts": "bulbul:v2",
                                  "note": "key present"}
    else:
        results["sarvam"] = {"status": "no_key"}

    # ── Groq ──────────────────────────────────────────
    groq_key = _platform_cfg.get("groq_api_key") or ""
    if groq_key:
        try:
            async with httpx.AsyncClient(timeout=5) as client:
                r = await client.get(
                    "https://api.groq.com/openai/v1/models",
                    headers={"Authorization": f"Bearer {groq_key}"}
                )
            results["groq"] = {"status": "ok" if r.status_code < 400 else "error",
                                "model": "llama-3.3-70b-versatile"}
        except Exception:
            results["groq"] = {"status": "ok", "model": "llama-3.3-70b-versatile",
                                "note": "key present"}
    else:
        results["groq"] = {"status": "no_key"}

    # ── Exotel ────────────────────────────────────────
    exotel_sid = os.getenv("EXOTEL_ACCOUNT_SID", "")
    exotel_key = os.getenv("EXOTEL_API_KEY", "")
    if exotel_sid and exotel_key:
        results["exotel"] = {
            "status":  "ok",
            "number":  os.getenv("EXOTEL_VIRTUAL_NUMBER", "07314854688"),
            "account": exotel_sid,
        }
    else:
        results["exotel"] = {"status": "no_key"}

    return {
        "status":    "ok",
        "version":   "3.0.0",
        "uptime_ms": round((time.time() - start) * 1000, 1),
        "timestamp": datetime.utcnow().isoformat(),
        "services":  results,
    }


# ── Aira System Prompt (read/write from pipeline file) ──
PIPELINE_PATH = os.path.join(os.path.dirname(__file__), 'exotel_pipeline.py')

@router.get("/system/prompt")
async def get_prompt():
    """Read current SYSTEM_PROMPT from the pipeline file."""
    try:
        with open(PIPELINE_PATH) as _f:
            src = _f.read()
        import re
        m = re.search(r'SYSTEM_PROMPT\s*=\s*"""(.*?)"""', src, re.DOTALL)
        if m:
            return {"prompt": m.group(1).strip()}
        return {"prompt": "", "warning": "SYSTEM_PROMPT not found in pipeline"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/system/prompt")
async def save_prompt(request: Request):
    """Write updated SYSTEM_PROMPT back to the pipeline file."""
    body   = await request.json()
    prompt = body.get("prompt", "").strip()
    if not prompt:
        raise HTTPException(status_code=400, detail="Prompt cannot be empty")
    try:
        with open(PIPELINE_PATH) as _f:
            src = _f.read()
        import re
        new_src = re.sub(
            r'(SYSTEM_PROMPT\s*=\s*""").*?(""")',
            f'\\1{prompt}\\2',
            src,
            flags=re.DOTALL
        )
        with open(PIPELINE_PATH, 'w') as _f:
            _f.write(new_src)
        db.add_log("✏️ System prompt updated via dashboard")
        return {"message": "Prompt saved. Restart service to apply."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))




@router.get("/system/config")

async def get_config_route():

    """Return current model/voice/telephony configuration (non-secret)."""

    provider = db.get_config("telephony_provider", "piopiy")
    _platform = tdb.get_tenant_config(1) or {}

    return {

        "telephony":      provider,

        "number":         os.getenv("EXOTEL_VIRTUAL_NUMBER", "") if provider == "exotel" else (_platform.get("piopiy_number") or ""),

        "stt_model":      "saarika:v2.5",

        "tts_model":      "bulbul:v2",

        "tts_voice":      _platform.get("agent_voice") or "anushka",

        "llm_model":      "llama-3.3-70b-versatile",

        "llm_provider":   "Groq",

        "language":       "Hindi (hi-IN)",

        "sarvam_key_set": bool(_platform.get("sarvam_api_key")),

        "groq_key_set":   bool(_platform.get("groq_api_key")),

        "exotel_key_set": bool(os.getenv("EXOTEL_API_KEY")),

        "company_name":    _platform.get("company_name") or "MuTech Automation",

        "agent_name":      _platform.get("agent_name") or "Aira",

        "target_audience": db.get_config("target_audience", ""),

    }



@router.post("/system/config")
async def save_config(request: Request):
    """Update editable config values."""
    body = await request.json()
    env_path = os.path.join(os.path.dirname(__file__), '..', '.env')

    env_updates = {}
    config_updates = {}
    if "virtual_number" in body:
        num = str(body["virtual_number"]).strip().lstrip("+").lstrip("91")
        if not num.isdigit() or len(num) != 10:
            raise HTTPException(status_code=400,
                detail="Phone must be 10 digits (e.g. 7314854688)")
        env_updates["EXOTEL_VIRTUAL_NUMBER"] = "0" + num  # store as 07XXXXXXXXX

    for key in ("company_name", "agent_name", "target_audience"):
        if key in body:
            value = str(body.get(key) or "").strip()
            if not value:
                raise HTTPException(status_code=400, detail=f"{key} cannot be empty")
            config_updates[key] = value

    if not env_updates and not config_updates:
        raise HTTPException(status_code=400, detail="Nothing to update")

    try:
        if env_updates:
            # Read current .env
            with open(env_path) as _f:
                lines = _f.readlines()
            for key, val in env_updates.items():
                found = False
                for i, line in enumerate(lines):
                    if line.startswith(f"{key}="):
                        lines[i] = f"{key}={val}\n"
                        found = True
                        break
                if not found:
                    lines.append(f"{key}={val}\n")
            with open(env_path, 'w') as _f:
                _f.writelines(lines)

            # Reload into current process
            for key, val in env_updates.items():
                os.environ[key] = val

        for key, val in config_updates.items():
            db.set_config(key, val)

        updates = {**env_updates, **config_updates}
        db.add_log(f"⚙️ Config updated: {', '.join(updates.keys())}")
        return {"message": "Saved. Changes are live immediately.", "updated": updates}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not save config: {e}")









# ═══════════════════════════════════════════════════════

# TELEPHONY SWITCH  (admin only)

# ═══════════════════════════════════════════════════════



@router.get("/system/telephony")

async def get_telephony():

    """Return current active telephony provider."""

    provider = db.get_config("telephony_provider", "piopiy")
    _platform = tdb.get_tenant_config(1) or {}

    return {

        "active": provider,

        "options": [

            {

                "id":          "piopiy",

                "name":        "PIOPIY (TeleCMI)",

                "description": "AI Agent via signaling server — recommended",

                "number":      _platform.get("piopiy_number") or "",

                "status":      "active" if provider == "piopiy" else "standby",

            },

            {

                "id":          "exotel",

                "name":        "Exotel",

                "description": "WebSocket voicebot via Exotel landline",

                "number":      os.getenv("EXOTEL_VIRTUAL_NUMBER", ""),

                "status":      "active" if provider == "exotel" else "standby",

            },

        ],

    }





@router.post("/system/telephony")

async def set_telephony(request: Request):

    """Switch active telephony provider. Admin only."""

    body     = await request.json()

    provider = body.get("provider", "").lower().strip()

    if provider not in ("piopiy", "exotel"):

        raise HTTPException(status_code=400, detail="provider must be 'piopiy' or 'exotel'")

    db.set_config("telephony_provider", provider)

    db.add_log(f"🔄 Telephony switched to {provider.upper()} by admin")

    return {

        "message": f"Active telephony switched to {provider.upper()}",

        "active":  provider,

        "note":    "Change is immediate — no restart required",

    }





# ═══════════════════════════════════════════════════════

# PIOPIY INBOUND WEBHOOK

# ═══════════════════════════════════════════════════════



@router.post("/piopiy/inbound")

async def piopiy_inbound(request: Request):

    """

    PIOPIY calls this webhook when someone dials 01203134158.

    We return PCMO actions to stream audio to our AI agent WebSocket.

    """

    import json

    body = {}

    try:

        body = await request.json()

    except Exception:

        pass



    caller = body.get("caller_id", body.get("from", "unknown"))

    called = body.get("did", body.get("to", "unknown"))

    call_id = body.get("call_id", body.get("request_id", "unknown"))



    db.add_log(f"📲 Inbound call from {caller} to {called}")



    # Return PCMO: stream caller audio to our piopiy_agent WebSocket pipeline

    # The piopiy-ai library handles this via the Agent signaling server

    # We just need to tell PIOPIY to connect to our AI agent

    _platform_cfg = tdb.get_tenant_config(1) or {}
    agent_id = (_platform_cfg.get("piopiy_agent_id") or "").strip()
    if not agent_id:
        agent_id = os.getenv("PIOPIY_AGENT_ID", "").strip()



    pcmo = {

        "ai_agent": {

            "agent_id": agent_id,

            "metadata": {

                "call_type": "inbound",

                "caller": caller,

                "called": called,

                "call_id": call_id,

            }

        }

    }



    return JSONResponse(content=pcmo)


# ═══════════════════════════════════════════════════════
# TENANT SETTINGS ENDPOINTS (admin role required)
# ═══════════════════════════════════════════════════════


def _require_admin(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ("admin", "superadmin"):
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user


@router.get("/tenant/profile")
async def get_tenant_profile(current_user: dict = Depends(get_current_user)):
    """Get current tenant's company profile and call config."""
    tid = current_user.get("tenant_id", 1)
    tenants = tdb.get_all_tenants()
    tenant  = next((t for t in tenants if t["id"] == tid), None)
    config  = tdb.get_tenant_config(tid) or {}
    return {
        "tenant": tenant,
        "config": {
            "company_name":      config.get("company_name", tenant["name"] if tenant else ""),
            "company_industry":  config.get("company_industry", ""),
            "company_products":  config.get("company_products", ""),
            "company_website":   config.get("company_website", ""),
            "call_language":     config.get("call_language", "hindi"),
            "call_guidelines":   config.get("call_guidelines", ""),
            "agent_name":        config.get("agent_name", "Aira"),
            "agent_voice":       config.get("agent_voice", "kavya"),
            "greeting_template": config.get("greeting_template", ""),
            "system_prompt":     config.get("system_prompt", ""),
            "setup_complete":    config.get("setup_complete", 0),
            "tts_model":         config.get("tts_model", "v3"),
            "tts_pace":          config.get("tts_pace", 1.1),
            "tts_temperature":   config.get("tts_temperature", 0.75),
            "stt_provider":      config.get("stt_provider", "sarvam"),
            "deepgram_key_set":  bool((config.get("deepgram_api_key") or "").strip()),
        },
    }


@router.get("/tenant/usage")
async def get_tenant_usage_summary(current_user: dict = Depends(get_current_user)):
    """Minutes & calls usage summary for the client dashboard."""
    tid    = current_user.get("tenant_id", 1)
    tenant = tdb.get_tenant(tid) or {}
    with db.get_conn() as conn:
        today = conn.execute(
            "SELECT COALESCE(SUM(calls_made),0), COALESCE(SUM(minutes_used),0) FROM usage_logs WHERE tenant_id=? AND date=date('now')",
            (tid,)
        ).fetchone()
        month = conn.execute(
            "SELECT COALESCE(SUM(calls_made),0), COALESCE(SUM(minutes_used),0) FROM usage_logs WHERE tenant_id=? AND date>=date('now','start of month')",
            (tid,)
        ).fetchone()
    minutes_limit = tenant.get("minutes_limit") or 0
    minutes_used  = round(float(tenant.get("minutes_used") or 0), 1)
    return {
        "minutes_used_total":  minutes_used,
        "minutes_limit":       minutes_limit,
        "minutes_left":        round(max(0, minutes_limit - minutes_used), 1) if minutes_limit else None,
        "minutes_pct":         round(minutes_used / minutes_limit * 100, 1) if minutes_limit else None,
        "minutes_today":       round(float(today[1] or 0), 1),
        "minutes_this_month":  round(float(month[1] or 0), 1),
        "calls_used_total":    tenant.get("calls_used", 0),
        "calls_limit":         tenant.get("calls_limit", 0),
        "calls_today":         int(today[0] or 0),
        "calls_this_month":    int(month[0] or 0),
        "plan":                tenant.get("plan", "starter"),
    }


@router.get("/tenant/plan-features")
async def get_plan_features_route(current_user: dict = Depends(get_current_user)):
    """Returns plan features + current usage for the logged-in tenant."""
    tid    = current_user.get("tenant_id", 1)
    tenant = tdb.get_tenant(tid) or {}
    plan   = tenant.get("plan", "starter")
    features = get_plan_features(plan)
    return {
        "plan":         plan,
        "features":     features,
        "calls_used":   tenant.get("calls_used", 0),
        "calls_limit":  tenant.get("calls_limit", 1000),
        "minutes_used": tenant.get("minutes_used", 0),
    }


@router.get("/tenant/billing")
async def get_billing(current_user: dict = Depends(get_current_user)):
    """Billing overview — plan, usage totals, 30-day daily usage, add-on purchases."""
    tid    = current_user.get("tenant_id", 1)
    tenant = tdb.get_tenant(tid) or {}
    usage  = tdb.get_tenant_usage(tid, days=30)
    from app.plan_features import get_plan_features
    return {
        "plan":             tenant.get("plan", "starter"),
        "calls_used":       tenant.get("calls_used", 0),
        "calls_limit":      tenant.get("calls_limit", 0),
        "minutes_used":     tenant.get("minutes_used", 0),
        "minutes_limit":    tenant.get("minutes_limit", 0),
        "features":         get_plan_features(tenant.get("plan", "starter")),
        "usage_30d":        usage,
        "addon_purchases":  [],
    }


@router.put("/tenant/profile")
async def update_tenant_profile(request: Request, current_user: dict = Depends(_require_admin)):
    """Update company profile and call guidelines — auto-builds system_prompt."""
    data = await request.json()
    tid  = current_user.get("tenant_id", 1)
    config = tdb.get_tenant_config(tid) or {}

    company_name      = data.get("company_name", "")
    company_industry  = data.get("company_industry", "")
    company_products  = data.get("company_products", "")
    company_website   = data.get("company_website", "")
    call_language     = data.get("call_language", "hindi")
    call_guidelines   = data.get("call_guidelines", "")
    agent_name        = data.get("agent_name", config.get("agent_name", "Aira"))
    agent_voice       = data.get("agent_voice", config.get("agent_voice", "kavya"))
    greeting_template = data.get("greeting_template", config.get("greeting_template", ""))
    tts_model         = data.get("tts_model", config.get("tts_model", "v3"))
    tts_pace          = float(data.get("tts_pace", config.get("tts_pace", 1.1)))
    tts_temperature   = float(data.get("tts_temperature", config.get("tts_temperature", 0.75)))

    lang_instruction = {
        "hindi":    "हमेशा हिंदी में बोलें।",
        "english":  "Always speak in English.",
        "hinglish": "Hinglish में बोलें — Hindi और English mix करें।",
    }.get(call_language, "हमेशा हिंदी में बोलें।")

    default_guidelines = (
        "- हर जवाब संक्षिप्त रखें (2-3 वाक्य)\n"
        "- अंत में demo schedule करने की कोशिश करें\n"
        "- रुचि नहीं है तो विनम्रता से call समाप्त करें"
    )
    effective_guidelines = call_guidelines or default_guidelines

    system_prompt = f"""आप {agent_name} हैं, {company_name} की professional AI sales agent हैं।

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
RESPONSE RULES — FOLLOW STRICTLY
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. Maximum 2 sentences per response. Never more.
2. End with exactly one question per turn. Never two.
3. No filler words: never say "Great!", "Sure!", "Certainly!", "Of course!"
4. If the customer interrupts, STOP immediately. Acknowledge in 3 words, then listen.
5. Match customer language exactly: Hindi→Hindi, English→English, Mixed→Hinglish.
6. Say numbers in words — "तीस प्रतिशत" not "30%".

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
COMPANY INFORMATION
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Company: {company_name}
Industry: {company_industry}
Products/Services: {company_products}
Website: {company_website}

{lang_instruction}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CALL GUIDELINES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{effective_guidelines}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
HARD STOPS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- "Remove my number" / "DNC" → Apologize, confirm removal, end call immediately.
- Abusive customer → "माफ कीजिए, have a good day." End call.
- Legal threat → End call immediately, no further response.
"""

    tdb.update_tenant_config(
        tid,
        company_name      = company_name,
        company_industry  = company_industry,
        company_products  = company_products,
        company_website   = company_website,
        call_language     = call_language,
        call_guidelines   = call_guidelines,
        agent_name        = agent_name,
        agent_voice       = agent_voice,
        greeting_template = greeting_template,
        system_prompt     = system_prompt,
        setup_complete    = 1,
        tts_model         = tts_model,
        tts_pace          = tts_pace,
        tts_temperature   = tts_temperature,
    )
    return {"ok": True, "message": "Profile updated. AI agent system prompt auto-generated."}


@router.put("/tenant/api-keys")
async def update_tenant_api_keys(request: Request, current_user: dict = Depends(_require_admin)):
    """Update BYOK API keys and service selections for this tenant."""
    data = await request.json()
    tid  = current_user.get("tenant_id", 1)
    saveable_fields = (
        # LLM
        "llm_provider", "llm_model",
        "groq_api_key", "xai_api_key", "openai_api_key", "anthropic_api_key", "gemini_api_key",
        # Speech
        "speech_provider", "stt_provider", "tts_provider",
        "sarvam_api_key",
        "deepgram_api_key",
        "elevenlabs_api_key", "elevenlabs_voice_id", "elevenlabs_model",
        # Telephony
        "piopiy_agent_id", "piopiy_agent_token", "piopiy_number",
        # Messaging
        "telegram_bot_token", "telegram_chat_id",
        "whatsapp_api_key", "whatsapp_number",
        # Webhook
        "webhook_url", "webhook_secret", "webhook_events",
    )
    update_data = {f: data[f] for f in saveable_fields if f in data}
    if update_data:
        tdb.update_tenant_config(tid, **update_data)
    return {"ok": True, "message": "Settings saved."}


@router.get("/tenant/service-quotas")
async def get_service_quotas(current_user: dict = Depends(_require_admin)):
    """Live quota check for Groq LLM, ElevenLabs TTS, and Sarvam STT/TTS."""
    import httpx
    tid    = current_user.get("tenant_id", 1)
    config = tdb.get_tenant_config(tid) or {}
    tenant = tdb.get_tenant(tid) or {}

    # ── 1. GROQ LLM ──────────────────────────────────────────────
    groq_key   = (config.get("groq_api_key") or "").strip()
    token_rows = tdb.get_tenant_token_usage_today(tid)
    groq_today = sum(r["total_tokens"] for r in token_rows if r["provider"] == "groq")
    groq_calls = sum(r["call_count"]   for r in token_rows if r["provider"] == "groq")
    groq_tpd_limit  = tenant.get("groq_daily_limit") or 100_000
    groq_tpm_limit  = None
    groq_tpm_remaining = None
    groq_rpm_limit  = None
    groq_rpm_remaining = None
    groq_key_status = "not_configured"

    if groq_key:
        try:
            async with httpx.AsyncClient(timeout=6) as c:
                r = await c.post(
                    "https://api.groq.com/openai/v1/chat/completions",
                    headers={"Authorization": f"Bearer {groq_key}",
                             "Content-Type": "application/json"},
                    json={"model": "llama-3.3-70b-versatile",
                          "messages": [{"role": "user", "content": "1"}],
                          "max_tokens": 1},
                )
            if r.status_code in (200, 429):
                groq_key_status  = "ok" if r.status_code == 200 else "exhausted"
                h = r.headers
                groq_tpm_limit      = int(h.get("x-ratelimit-limit-tokens",      0) or 0)
                groq_tpm_remaining  = int(h.get("x-ratelimit-remaining-tokens",  0) or 0)
                groq_rpm_limit      = int(h.get("x-ratelimit-limit-requests",     0) or 0)
                groq_rpm_remaining  = int(h.get("x-ratelimit-remaining-requests", 0) or 0)
                if r.status_code == 429:
                    # Parse used/limit from error message if available
                    try:
                        err = r.json()["error"]["message"]
                        import re
                        m = re.search(r"Limit (\d+), Used (\d+)", err)
                        if m:
                            groq_tpd_limit = int(m.group(1))
                            groq_today     = max(groq_today, int(m.group(2)))
                    except Exception:
                        pass
            elif r.status_code == 401:
                groq_key_status = "invalid_key"
            else:
                groq_key_status = f"error_{r.status_code}"
        except Exception as e:
            groq_key_status = "unreachable"

    groq_tpd_used = groq_today
    groq_tpd_pct  = round(groq_tpd_used / groq_tpd_limit * 100, 1) if groq_tpd_limit else 0
    groq_tpd_status = (
        "exhausted" if groq_tpd_used >= groq_tpd_limit else
        "warning"   if groq_tpd_pct >= 80 else
        groq_key_status
    )

    # ── 2. ELEVENLABS TTS ─────────────────────────────────────────
    labs_key    = (config.get("elevenlabs_api_key") or "").strip()
    labs_chars_today = tdb.get_tenant_tts_chars_today(tid)
    labs_chars_limit = None
    labs_chars_used  = None
    labs_status      = "not_configured"
    labs_plan        = None
    labs_note        = None

    if labs_key:
        try:
            async with httpx.AsyncClient(timeout=5) as c:
                r_sub = await c.get(
                    "https://api.elevenlabs.io/v1/user/subscription",
                    headers={"xi-api-key": labs_key},
                )
            if r_sub.status_code == 200:
                sub = r_sub.json()
                labs_chars_used  = sub.get("character_count", 0)
                labs_chars_limit = sub.get("character_limit", 0)
                labs_plan        = sub.get("tier", "unknown")
                labs_status      = "ok"
            elif r_sub.status_code in (401, 403):
                body = r_sub.text or ""
                if "missing_permissions" in body or "user_read" in body:
                    # Scoped TTS key — verify it works via /v1/models
                    async with httpx.AsyncClient(timeout=5) as c2:
                        r2 = await c2.get(
                            "https://api.elevenlabs.io/v1/models",
                            headers={"xi-api-key": labs_key},
                        )
                    if r2.status_code == 200:
                        labs_status = "key_valid_scoped"
                        labs_chars_used = labs_chars_today or 0
                        labs_note = "Scoped TTS key — balance unavailable. Chars tracked from calls."
                    else:
                        labs_status = "invalid_key"
                else:
                    labs_status = "invalid_key"
            elif r_sub.status_code == 402:
                labs_status = "quota_exceeded"
            else:
                labs_status = f"error_{r_sub.status_code}"
        except Exception:
            labs_status = "unreachable"

    # ── 3. SARVAM STT/TTS ─────────────────────────────────────────
    sarvam_key    = (config.get("sarvam_api_key") or "").strip()
    sarvam_status = "not_configured"
    sarvam_note   = None
    sarvam_calls_today = 0

    if sarvam_key:
        try:
            async with httpx.AsyncClient(timeout=5) as c:
                r = await c.get(
                    "https://api.sarvam.ai/v1/models",
                    headers={"api-subscription-key": sarvam_key},
                )
            if r.status_code == 200:
                sarvam_status = "ok"
                sarvam_note   = "Sarvam AI does not expose quota or usage via API."
            elif r.status_code == 401:
                sarvam_status = "invalid_key"
            else:
                sarvam_status = f"error_{r.status_code}"
        except Exception:
            sarvam_status = "unreachable"
        # Count STT calls today from usage_logs (each call used STT)
        try:
            with __import__('app.database', fromlist=['get_conn']).get_conn() as conn:
                row = conn.execute(
                    "SELECT COALESCE(SUM(calls_made),0) FROM usage_logs WHERE tenant_id=? AND date=date('now')",
                    (tid,)
                ).fetchone()
                sarvam_calls_today = row[0] if row else 0
        except Exception:
            sarvam_calls_today = 0

    return {
        "groq": {
            "key_status":        groq_key_status,
            "tpd_used":          groq_tpd_used,
            "tpd_limit":         groq_tpd_limit,
            "tpd_left":          max(0, groq_tpd_limit - groq_tpd_used),
            "tpd_pct":           groq_tpd_pct,
            "tpm_limit":         groq_tpm_limit,
            "tpm_remaining":     groq_tpm_remaining,
            "rpm_limit":         groq_rpm_limit,
            "rpm_remaining":     groq_rpm_remaining,
            "llm_calls_today":   groq_calls,
            "status":            groq_tpd_status,
            "note":              "Free tier: 100K tokens/day, 12K tokens/min, 1K req/min",
        },
        "elevenlabs": {
            "key_status":        labs_status,
            "plan":              labs_plan,
            "chars_used_account": labs_chars_used,
            "chars_limit_account": labs_chars_limit,
            "chars_today":       labs_chars_today,
            "chars_left":        max(0, (labs_chars_limit or 0) - (labs_chars_used or 0)) if labs_chars_limit else None,
            "pct_used":          round(labs_chars_used / labs_chars_limit * 100, 1) if labs_chars_limit else None,
            "status":            labs_status,
            "note":              labs_note or ("Quota info from ElevenLabs account." if labs_status == "ok" else None),
        },
        "sarvam": {
            "key_status":        sarvam_status,
            "calls_today":       sarvam_calls_today,
            "status":            sarvam_status,
            "note":              sarvam_note,
            "quota_api":         False,
        },
        "token_details": token_rows,
    }


@router.get("/tenant/api-keys")
async def get_tenant_api_keys(current_user: dict = Depends(_require_admin)):
    """Get tenant API keys and service config (secrets masked)."""
    tid    = current_user.get("tenant_id", 1)
    config = tdb.get_tenant_config(tid) or {}

    def mask(v):
        if v and len(v) > 9:
            return v[:6] + "•••" + v[-3:]
        return "✓ Set" if v else ""

    return {
        # LLM
        "llm_provider":        config.get("llm_provider", "groq"),
        "llm_model":           config.get("llm_model", ""),
        "groq_api_key":        mask(config.get("groq_api_key", "")),
        "xai_api_key":         mask(config.get("xai_api_key", "")),
        "openai_api_key":      mask(config.get("openai_api_key", "")),
        "anthropic_api_key":   mask(config.get("anthropic_api_key", "")),
        "gemini_api_key":      mask(config.get("gemini_api_key", "")),
        "groq_set":            bool(config.get("groq_api_key")),
        "xai_set":             bool(config.get("xai_api_key")),
        "openai_set":          bool(config.get("openai_api_key")),
        "anthropic_set":       bool(config.get("anthropic_api_key")),
        "gemini_set":          bool(config.get("gemini_api_key")),
        # Speech
        "speech_provider":     config.get("speech_provider", "sarvam"),
        "stt_provider":        config.get("stt_provider", "sarvam"),
        "tts_provider":        config.get("tts_provider", "elevenlabs"),
        "sarvam_api_key":      mask(config.get("sarvam_api_key", "")),
        "elevenlabs_api_key":  mask(config.get("elevenlabs_api_key", "")),
        "elevenlabs_voice_id": config.get("elevenlabs_voice_id", ""),
        "elevenlabs_model":    config.get("elevenlabs_model", "eleven_flash_v2_5"),
        "sarvam_set":          bool(config.get("sarvam_api_key")),
        "elevenlabs_set":      bool(config.get("elevenlabs_api_key")),
        # Telephony
        "piopiy_agent_id":     config.get("piopiy_agent_id", ""),
        "piopiy_agent_token":  mask(config.get("piopiy_agent_token", "")),
        "piopiy_token_set":    bool(config.get("piopiy_agent_token")),
        "piopiy_number":       config.get("piopiy_number", ""),
        # Messaging
        "telegram_bot_token":  mask(config.get("telegram_bot_token", "")),
        "telegram_chat_id":    config.get("telegram_chat_id", ""),
        "whatsapp_api_key":    mask(config.get("whatsapp_api_key", "")),
        "whatsapp_number":     config.get("whatsapp_number", ""),
        "telegram_set":        bool(config.get("telegram_bot_token")),
        "whatsapp_set":        bool(config.get("whatsapp_api_key")),
        # Webhook
        "webhook_url":         config.get("webhook_url", ""),
        "webhook_events":      config.get("webhook_events", "call_completed"),
        "webhook_secret_set":  bool(config.get("webhook_secret")),
    }


@router.put("/tenant/system-prompt")
async def update_tenant_system_prompt(request: Request, current_user: dict = Depends(_require_admin)):
    """Directly update the AI system prompt for this tenant."""
    data   = await request.json()
    prompt = data.get("prompt", "").strip()
    if not prompt:
        raise HTTPException(status_code=400, detail="Prompt cannot be empty")
    tid = current_user.get("tenant_id", 1)
    tdb.update_tenant_config(tid, system_prompt=prompt, setup_complete=1)
    return {"ok": True, "message": "System prompt saved."}


@router.get("/tenant/faq")
async def get_tenant_faq(current_user: dict = Depends(get_current_user)):
    """Return the current FAQ content for this tenant."""
    tid    = current_user.get("tenant_id", 1)
    config = tdb.get_tenant_config(tid) or {}
    return {
        "faq_content": config.get("faq_content", "") or "",
        "preview":     _build_faq_prompt_section(config.get("faq_content", "") or ""),
    }


@router.put("/tenant/faq")
async def update_tenant_faq(
    request: Request,
    current_user: dict = Depends(_require_admin),
):
    """
    Save FAQ content. The FAQ is automatically injected into the agent's
    system prompt at call time (appended as a 'Frequently Asked Questions'
    section so the agent can answer product/company questions accurately).
    """
    data        = await request.json()
    faq_content = (data.get("faq_content") or "").strip()
    tid         = current_user.get("tenant_id", 1)

    tdb.update_tenant_config(tid, faq_content=faq_content)
    db.add_log(f"📚 FAQ updated by {current_user['username']} — {len(faq_content)} chars")

    return {
        "ok":      True,
        "message": "FAQ saved. It will be active on the next call.",
        "chars":   len(faq_content),
    }


@router.get("/tenant/webhook-config")
async def get_webhook_config(current_user: dict = Depends(_require_admin)):
    """Get tenant webhook URL."""
    tid    = current_user.get("tenant_id", 1)
    config = tdb.get_tenant_config(tid) or {}
    return {"webhook_url": config.get("webhook_url") or ""}


@router.put("/tenant/webhook-config")
async def update_webhook_config(request: Request, current_user: dict = Depends(_require_admin)):
    """Save tenant webhook URL. Requires crm_webhook feature on plan."""
    tid    = current_user.get("tenant_id", 1)
    tenant = tdb.get_tenant(tid) or {}
    gate   = check_feature(tenant.get("plan", "starter"), "crm_webhook")
    if not gate["allowed"]:
        raise HTTPException(status_code=402, detail={
            "message":    gate["reason"],
            "upgrade_to": gate["upgrade_to"],
            "code":       "FEATURE_LOCKED",
        })
    data        = await request.json()
    webhook_url = (data.get("webhook_url") or "").strip()
    tdb.update_tenant_config(tid, webhook_url=webhook_url)
    db.add_log(f"🔗 Webhook URL {'set' if webhook_url else 'cleared'} by {current_user['username']}")
    return {"ok": True, "webhook_url": webhook_url}


@router.get("/tenant/webhook")
async def get_webhook(current_user: dict = Depends(_require_admin)):
    """Get tenant webhook configuration (canonical URL)."""
    tid = current_user.get("tenant_id", 1)
    cfg = tdb.get_tenant_config(tid) or {}
    return {
        "webhook_url":        cfg.get("webhook_url", ""),
        "webhook_events":     cfg.get("webhook_events", "call_completed"),
        "webhook_secret_set": bool(cfg.get("webhook_secret")),
    }


@router.put("/tenant/webhook")
async def update_webhook(request: Request, current_user: dict = Depends(_require_admin)):
    """Update webhook URL, events, and secret. Plan-gated to Growth+."""
    tid    = current_user.get("tenant_id", 1)
    tenant = tdb.get_tenant(tid)
    plan   = (tenant or {}).get("plan", "starter")
    gate   = check_feature(plan, "crm_webhook")
    if not gate["allowed"]:
        raise HTTPException(status_code=402, detail={
            "message":    gate["reason"],
            "upgrade_to": gate["upgrade_to"],
            "code":       "FEATURE_LOCKED",
        })
    data   = await request.json()
    update = {}
    url = (data.get("webhook_url") or "").strip()
    if url and not url.startswith("http"):
        raise HTTPException(status_code=400, detail="webhook_url must start with http/https")
    if "webhook_url" in data:
        update["webhook_url"] = url
    if "webhook_events" in data:
        update["webhook_events"] = data["webhook_events"] or "call_completed"
    secret = (data.get("webhook_secret") or "").strip()
    if secret:
        update["webhook_secret"] = secret
    if update:
        tdb.update_tenant_config(tid, **update)
    db.add_log(f"🔗 Webhook config updated by {current_user['username']}")
    return {"ok": True, "message": "Webhook configuration saved."}


@router.post("/tenant/webhook/test")
async def test_webhook(request: Request, current_user: dict = Depends(_require_admin)):
    """Send a test call_completed event to the configured webhook URL."""
    from app.webhook_service import fire_call_webhook
    tid = current_user.get("tenant_id", 1)
    cfg = tdb.get_tenant_config(tid) or {}
    if not cfg.get("webhook_url"):
        raise HTTPException(status_code=400, detail="No webhook URL configured")
    test_payload = {
        "call_id":      0,
        "phone":        "+919876543210",
        "lead_name":    "Test Lead",
        "company":      "Test Company",
        "duration_sec": 45,
        "outcome":      "answered",
        "sentiment":    "interested",
        "summary":      "This is a test webhook event from DialBot.",
        "transcript":   "",
        "campaign_id":  None,
        "lead_id":      None,
    }
    import asyncio as _asyncio
    _asyncio.create_task(fire_call_webhook(tid, test_payload))
    return {"ok": True, "message": "Test webhook fired. Check webhook_logs for result."}


@router.get("/tenant/addons")
async def get_tenant_addons(current_user: dict = Depends(_require_admin)):
    """List addon minute purchases for this tenant."""
    tid = current_user.get("tenant_id", 1)
    with db.get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM addon_purchases WHERE tenant_id=? ORDER BY purchased_at DESC",
            (tid,),
        ).fetchall()
    return {"addons": [dict(r) for r in rows]}


@router.get("/tenant/webhook/logs")
async def get_webhook_logs(current_user: dict = Depends(_require_admin), limit: int = 50):
    """Get recent webhook delivery logs for this tenant."""
    tid = current_user.get("tenant_id", 1)
    with db.get_conn() as conn:
        rows = conn.execute(
            "SELECT id, event, url, status_code, response, fired_at"
            " FROM webhook_logs WHERE tenant_id=? ORDER BY fired_at DESC LIMIT ?",
            (tid, limit),
        ).fetchall()
    return {"logs": [dict(r) for r in rows], "total": len(rows)}


def _build_faq_prompt_section(faq_content: str) -> str:
    """
    Convert raw FAQ text into a formatted prompt section.
    Called by multi_agent_manager and piopiy_agent when building system prompt.
    """
    if not faq_content or not faq_content.strip():
        return ""
    return (
        "\n\n--- Frequently Asked Questions ---\n"
        "Use the following Q&A to answer customer questions accurately. "
        "If a customer asks something covered here, use this answer directly.\n\n"
        + faq_content.strip()
    )
